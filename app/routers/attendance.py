"""
Attendance routes: login, logout, today status, history
"""
from typing import Optional
from fastapi import APIRouter, Depends, UploadFile, File, Form, status
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.user import User
from app.models.attendance import AttendanceStatus
from app.schemas.attendance import (
    AttendanceActionResponse,
    AttendanceHistoryResponse,
    TodayStatusResponse,
)
from app.services import attendance_service
from app.utils.dependencies import get_verified_user
from io import BytesIO
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date, timedelta
from calendar import monthrange

router = APIRouter(prefix="/api/v1/attendance", tags=["Attendance"])


@router.post(
    "/login",
    response_model=AttendanceActionResponse,
    status_code=status.HTTP_201_CREATED,
)
async def attendance_login(
    file: UploadFile = File(..., description="Selfie image"),
    latitude: float = Form(..., description="Current GPS latitude"),
    longitude: float = Form(..., description="Current GPS longitude"),
    address: Optional[str] = Form(None, description="Optional readable address"),
    current_user: User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    """
    Mark daily LOGIN attendance.
    Requires: selfie + GPS location.
    Face is verified against reference selfie uploaded at signup.
    """
    attendance = await attendance_service.mark_login(
        db, current_user, file, latitude, longitude, address
    )
    return AttendanceActionResponse(
        message="Login attendance marked successfully! Have a great workday 🎉",
        attendance=attendance,
    )


@router.post("/logout", response_model=AttendanceActionResponse)
async def attendance_logout(
    file: UploadFile = File(..., description="Selfie image"),
    latitude: float = Form(...),
    longitude: float = Form(...),
    address: Optional[str] = Form(None),
    current_user: User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    """
    Mark daily LOGOUT attendance.
    Must have an active login for today.
    """
    attendance = await attendance_service.mark_logout(
        db, current_user, file, latitude, longitude, address
    )
    return AttendanceActionResponse(
        message="Logout attendance marked successfully! See you tomorrow 👋",
        attendance=attendance,
    )


@router.get("/today", response_model=TodayStatusResponse)
def get_today_status(
    current_user: User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    """Get today's attendance status for current user"""
    attendance = attendance_service.get_today_attendance(db, current_user.id)

    if not attendance:
        return TodayStatusResponse(
            date=date.today(),
            is_logged_in=False,
            attendance=None,
            message="No attendance marked yet today. Please login.",
        )

    is_active = attendance.status == AttendanceStatus.ACTIVE
    msg = (
        "You are currently logged in. Don't forget to logout by 10 PM."
        if is_active
        else f"Today's attendance: {attendance.status.value}."
    )
    return TodayStatusResponse(
        date=attendance.date,
        is_logged_in=is_active,
        attendance=attendance,
        message=msg,
    )


@router.get("/history", response_model=AttendanceHistoryResponse)
def get_history(
    limit: int = 200,
    current_user: User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    """
    Get attendance history including LEAVES (weekdays with no attendance).
    Skips weekends. Returns records + stats.
    """
    records = attendance_service.get_attendance_with_leaves(db, current_user, limit)
    stats = attendance_service.get_leave_stats(records)
    return AttendanceHistoryResponse(
        total_days=stats["total_days"],
        full_days=stats["full_days"],
        half_days=stats["half_days"],
        active_days=stats["active_days"],
        leaves=stats["leaves"],
        records=records,
    )


@router.get("/export-excel")
def export_attendance_excel(
    current_user: User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    """
    Export user's attendance history as beautifully formatted Excel file.
    """
    records = attendance_service.get_attendance_with_leaves(db, current_user, limit=1000)
    stats = attendance_service.get_leave_stats(records)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Attendance Report — {current_user.full_name}"
    ws["A1"].font = Font(bold=True, size=16, color="1E293B")
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30

    # Meta info row
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Email: {current_user.email}   |   Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = Font(italic=True, size=10, color="64748B")
    ws["A2"].alignment = center

    # Stats row
    ws.merge_cells("A3:H3")
    ws["A3"] = (
        f"Total Days: {stats['total_days']}   |   "
        f"Full Days: {stats['full_days']}   |   "
        f"Half Days: {stats['half_days']}   |   "
        f"Leaves: {stats['leaves']}   |   "
        f"Active: {stats['active_days']}"
    )
    ws["A3"].font = Font(bold=True, size=11, color="4F46E5")
    ws["A3"].alignment = center
    ws.row_dimensions[3].height = 22

    # Empty row for spacing
    ws.row_dimensions[4].height = 8

    # Column headers (row 5)
    headers = [
        "S.No", "Date", "Day", "Login Time", "Logout Time",
        "Duration", "Status", "Location"
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[5].height = 28

    # Data rows
    for idx, r in enumerate(records, start=1):
        row_num = idx + 5
        date_obj = r["date"]
        date_str = date_obj.strftime("%d %b %Y") if date_obj else "—"
        day_str = date_obj.strftime("%A") if date_obj else "—"

        login_t = r["login_time"].strftime("%I:%M %p") if r["login_time"] else "—"
        logout_t = r["logout_time"].strftime("%I:%M %p") if r["logout_time"] else "—"

        # Duration
        duration = "—"
        if r["login_time"] and r["logout_time"]:
            diff = (r["logout_time"] - r["login_time"]).total_seconds() / 60
            h = int(diff // 60)
            m = int(diff % 60)
            duration = f"{h}h {m}m"

        # Status with emoji
        status_map = {
            "full_day": "🟢 Full Day",
            "half_day": "🟡 Half Day",
            "active":   "🔵 Active",
            "leave":    "🔴 Leave",
        }
        status_str = status_map.get(r["status"], r["status"])

        # Location
        location = r["login_address"] or "—"
        if r["status"] == "leave":
            location = "—"

        row_data = [idx, date_str, day_str, login_t, logout_t, duration, status_str, location]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = center if col_idx != 8 else Alignment(vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(size=11)

            # Color code the status column
            if col_idx == 7:
                status_key = r["status"]
                if status_key == "full_day":
                    cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                elif status_key == "half_day":
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                elif status_key == "leave":
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif status_key == "active":
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    # Auto column widths
    column_widths = [6, 15, 12, 14, 14, 12, 14, 50]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Freeze header row
    ws.freeze_panes = "A6"

    # Footer note
    footer_row = 6 + len(records) + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    footer_cell = ws.cell(row=footer_row, column=1, value="Generated by Attendify — Smart Attendance System")
    footer_cell.font = Font(italic=True, size=9, color="94A3B8")
    footer_cell.alignment = center

    # Save to memory buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Attendance_{current_user.full_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/salary")
def get_salary_report(
    month: int = None,
    year: int = None,
    current_user: User = Depends(get_verified_user),
    db: Session = Depends(get_db),
):
    """
    Get monthly salary breakdown with deductions based on attendance.
    Base salary: ₹30,000/month
    Deduction rules:
      - Full leave day: Full day salary deducted
      - Half day: Half day salary deducted
    """
    from calendar import monthrange
    from app.config import settings as app_settings
    from app.models.attendance import Attendance, AttendanceStatus

    # Default to current month if not provided
    today = date.today()
    if not month or not year:
        month = today.month
        year = today.year

    # Validate
    if month < 1 or month > 12:
        raise HTTPException(status_code=400, detail="Invalid month (1-12)")
    if year < 2020 or year > today.year + 1:
        raise HTTPException(status_code=400, detail="Invalid year")

    # Calculate month boundaries
    first_day = date(year, month, 1)
    last_day_num = monthrange(year, month)[1]
    last_day = date(year, month, last_day_num)

    # Count working days in month (Mon-Fri only)
    working_days = 0
    all_dates_in_month = []
    for day_num in range(1, last_day_num + 1):
        d = date(year, month, day_num)
        if d.weekday() < 5:  # Monday=0, Friday=4
            working_days += 1
        all_dates_in_month.append(d)

    if working_days == 0:
        raise HTTPException(status_code=400, detail="No working days in this month")

    # Per day salary
    monthly_salary = app_settings.MONTHLY_SALARY
    per_day_salary = round(monthly_salary / working_days, 2)

    # Get attendance records for this month
    records = db.query(Attendance).filter(
        Attendance.user_id == current_user.id,
        Attendance.date >= first_day,
        Attendance.date <= last_day,
    ).all()

    record_map = {r.date: r for r in records}

    # Only consider working days that have passed (up to today)
    # For current month, don't count future dates as leaves
    end_check_date = min(last_day, today - timedelta(days=1))
    if month == today.month and year == today.year:
        # For current month, only count days that have passed
        elapsed_working_days = 0
        for day_num in range(1, last_day_num + 1):
            d = date(year, month, day_num)
            if d.weekday() < 5 and d <= end_check_date:
                elapsed_working_days += 1
    else:
        # For past months, count all working days
        elapsed_working_days = working_days

    # Calculate deductions
    full_days = 0
    half_days = 0
    active_days = 0
    leaves = 0
    daily_breakdown = []

    for day_num in range(1, last_day_num + 1):
        d = date(year, month, day_num)
        is_weekend = d.weekday() >= 5
        is_future = d > end_check_date
        record = record_map.get(d)

        day_info = {
            "date": d.isoformat(),
            "day_name": d.strftime("%A"),
            "is_weekend": is_weekend,
            "is_future": is_future,
            "status": None,
            "deduction": 0.0,
        }

        if is_weekend:
            day_info["status"] = "weekend"
        elif is_future:
            day_info["status"] = "upcoming"
        elif record:
            status_str = record.status.value if hasattr(record.status, 'value') else str(record.status)
            day_info["status"] = status_str
            if status_str == "full_day":
                full_days += 1
            elif status_str == "half_day":
                half_days += 1
                day_info["deduction"] = round(per_day_salary * 0.5, 2)
            elif status_str == "active":
                active_days += 1
        else:
            # Working day but no attendance = leave
            day_info["status"] = "leave"
            leaves += 1
            day_info["deduction"] = per_day_salary

        daily_breakdown.append(day_info)

    # Total deductions
    leave_deduction = round(leaves * per_day_salary, 2)
    half_day_deduction = round(half_days * per_day_salary * 0.5, 2)
    total_deduction = round(leave_deduction + half_day_deduction, 2)
    final_salary = round(monthly_salary - total_deduction, 2)

    # Attendance rate
    attended = full_days + (half_days * 0.5)
    attendance_rate = round((attended / elapsed_working_days) * 100, 1) if elapsed_working_days else 0

    return {
        "month": month,
        "year": year,
        "month_name": first_day.strftime("%B %Y"),
        "currency": app_settings.CURRENCY,
        "monthly_salary": monthly_salary,
        "working_days_in_month": working_days,
        "elapsed_working_days": elapsed_working_days,
        "per_day_salary": per_day_salary,
        "summary": {
            "full_days": full_days,
            "half_days": half_days,
            "active_days": active_days,
            "leaves": leaves,
            "attendance_rate": attendance_rate,
        },
        "deductions": {
            "leave_deduction": leave_deduction,
            "half_day_deduction": half_day_deduction,
            "total_deduction": total_deduction,
        },
        "final_salary": final_salary,
        "daily_breakdown": daily_breakdown,
    }